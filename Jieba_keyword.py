# File Name: Jieba_keyword.py
# Date: 2022/4/3 
# Author: HJL
# IDE: PyCharm
import jieba.analyse
import codecs
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
import time


def main():
    start = time.time()
    path = r'C:\Users\xxx\Desktop\test.txt'
    # keywords1 = keyword_tfidf(path)
    keywords2 = keyword_textrank(path)
    # save_to_xlsx(keywords1, keywords2, path)
    end = time.time()
    print(f'Running time:\n{end-start} secs')


def keyword_tfidf(path):
    tfidf = jieba.analyse.extract_tags
    text = codecs.open(path, 'r', 'utf-8').read()
    comp = re.compile('[^A-Za-z0-9\u4e00-\u9fa5]')
    c_text = comp.sub('', text)
    keywords1 = tfidf(c_text, topK=20, withWeight=True, allowPOS=(), withFlag=False)
    print("keywords by tfidf:")
    for keyword, weight in keywords1:  # extract kw
        print(f'{keyword} {weight}')  # f = format 'xx' as string
    return keywords1


def keyword_textrank(path):
    textrank = jieba.analyse.textrank  # connect textrank
    print("keywords by textrank:")
    text = codecs.open(path, 'r', 'utf-8').read()
    comp = re.compile('[^A-Za-z0-9\u4e00-\u9fa5]')
    c_text = comp.sub('', text)
    keywords2 = textrank(c_text, topK=20, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v'))  # based on TextRank
    for keyword, weight in keywords2:  # extract kw
        print(f'{keyword} {weight}')
    return keywords2


def save_to_xlsx(keywords1, keywords2, path):
    wb = Workbook()
    wb.active.title = "TFIDF"  # active sheet rename
    wb.create_sheet("TEXTRANK")
    file = path + ".xlsx"
    wb.save(file)

    wb = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    # writer.book = wb # 设置不覆盖原有sheet

    df1 = pd.DataFrame(keywords1, columns=['Keyword', 'Weight'])
    df1.to_excel(writer, sheet_name='TFIDF', index=None)  # sheet_name='TFIDF'
    df2 = pd.DataFrame(keywords2, columns=['Keyword', 'Weight'])
    df2.to_excel(writer, sheet_name='TEXTRANK', index=None)  # sheet_name='TEXTRANK'

    writer.save()


if __name__ == "__main__":
    main()


"""tags = jieba.analyse.extract_tags(file, topK=20, withWeight=False, allowPOS=('ns', 'n', 'vn', 'v'), withFlag=False)
file 为待提取的文本，topK 为返回几个 TF/IDF 权重最大的关键词，默认值为 20
withWeight 为是否一并返回关键词权重值，默认值为 False，allowPOS 仅包括指定词性的词，tfidf默认不筛选, textrank默认筛选
withFlag=False 不同时输出词性"""

'''基于TF-IDF的关键词抽取算法:
目标是获取文本中词频高，也就是TF大的，且语料库其他文本中词频低的，
也就是IDF大的。这样的词可以作为文本的标志，用于区分其他文本'''

'''基于TextRank的关键词抽取算法:
1. 将文本进行分词和词性标注，将特定词性的词（比如名词）作为节点添加到图中。
2. 出现在一个窗口中的词语之间形成一条边，窗口大小可设置为2~10之间，它表示一个窗口中有多少个词语。
3. 对节点根据入度节点个数以及入度节点权重进行打分，入度节点越多，且入度节点权重大，则打分高。
4. 根据打分进行降序排列，输出指定个数的关键词。'''
